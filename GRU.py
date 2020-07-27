import numpy as np 
import pickle
import pandas as pd 
import datetime as dt
import matplotlib.pyplot as plt
import openpyxl
from scipy.stats import linregress
from sklearn.preprocessing import MinMaxScaler
np.random.seed(1000)
from keras.models import Sequential
from keras.layers import Dense
from keras.layers import GRU
from keras.layers import Dropout


def momentum(Close):
    returns = np.log(Close)
    x = np.arange(len(returns))
    slope, _, rvalue, _, _ = linregress(x, returns)
    return ((1 + slope) ** 252) * (rvalue ** 2)

def Tech_Ind_Ovrnightret(df):
    #Volatility
    
    df['daily_returns']=  df['PX_LAST'].pct_change()
    df['Volatility21d']= df['daily_returns'].rolling(21).std()
    #df = df.drop("daily_returns", axis =1)
    
    #Create 21 days Moving Average
    df['SMA21'] = df['PX_LAST'].rolling(window=21).mean()

    #Create Bollinger Bands()
    df['21sd'] = df['PX_LAST'].rolling(21).std()
    df['upper_band'] = df['SMA21'] + (df['21sd']*2)
    df['lower_band'] = df['SMA21'] - (df['21sd']*2)
    df = df.drop("21sd", axis =1)

    #Momentum
    df['Momentum'] = df['PX_LAST'].rolling(21).apply(momentum, raw=False)

    #Overnight returns
    df["Overnight_Return"] = df['PX_OPEN']/df['PX_LAST'].shift(periods=1)-1

    return df.loc[21:,:]

def monthly_change(feature):
    return feature.iloc[-1] / feature.iloc[0] - 1

# Training set 
def train_inputs_targets(df):
    #*** Preparing the data*****
    df['month'] = df['Dates'].astype(str).str[:7]
    df.index = df['month'] # the months are assigned as index
    df = df.drop(['month'], axis=1) # remove the months column 
    df = df.drop(['Dates'],axis=1) # remove the dates column
    df_train = df.loc[:training_stop_date,:] # truncate it to what we are interested in only
    df_train_scaled = scaler.fit_transform(df_train) # normalized the data
    df_train_scaled = pd.DataFrame(data = df_train_scaled) # transform it into dataframe 
    df_train_scaled.columns =  df_train.columns # get the columns of the df_scaled
    df_train_scaled.index = df_train.index # get the index of the df_scaled
    #**** y_train *****
    grouped_train = df_train.groupby("month") # we grouped by to perform an operation
    y_train = grouped_train['PX_LAST'].apply(monthly_change) # monthly returns of the training set
    y_train = y_train[1:] #we shift to +1month as we need the prediction of the subsequent month 
    y_train = np.array(y_train)
    y_train = np.reshape(y_train,(len(y_train),1)) # we transform it into a 2D
    #***** X_train *****
    list_features =list(df_train.columns) # List of features 
    list_months= list(df_train.index.drop_duplicates()) # List of months in the training set 
    X_train=[]
    t_steps= 21 #Number of days (data) per month 
    count= 0    # counter
    for month in list_months:    
        features =[]  #All column features
        for i in list_features:
            feat = list(df_train_scaled.loc[month,i]) #A single column feature
            if len(feat)< t_steps: # add data from previous month if there is not enough data (21)
                difference = t_steps - len(feat)
                feat_moins1 = list(df_train_scaled.loc[list_months[count-1],i])
                to_add= feat_moins1[-difference:]
                feat = to_add + feat
            elif len(feat)> t_steps: # truncate if there are too many data in one instance
                difference = len(feat) - t_steps
                feat = feat[difference:]
            features.append(feat)
        features = np.transpose(features)
        X_train.append(features)
        count += 1
    X_train = np.array(X_train)
    X_train =X_train[:-1,:,:]  #we shift to -1month as this is the predictor
    return X_train, y_train

# Testing set
def test_inputs_targets(df):
    #*** Preparing the data*****
    df['month'] = df['Dates'].astype(str).str[:7]
    df.index = df['month'] # the months are assigned as index
    df = df.drop(['month'], axis=1) # remove the months column 
    df = df.drop(['Dates'],axis=1) # remove the dates column
    df_test = df.loc[testing_start_date:,:] # truncate it to what we are interested in only
    df_test.to_csv('.\Brouillons\df_test2.csv')
    df_test_scaled = scaler.fit_transform(df_test) # normalized the data
    df_test_scaled = pd.DataFrame(data= df_test_scaled) # transform it into dataframe 
    df_test_scaled.columns =  df_test.columns # get the columns of the df_scaled
    df_test_scaled.index = df_test.index # get the index of the df_scaled
    df_test_scaled.to_csv('.\Brouillons\df_test.csv')
    #**** y_test *****
    grouped_test = df_test.groupby("month") # we grouped by to perform an operation
    y_test = grouped_test['PX_LAST'].apply(monthly_change) # monthly returns of the test set
    y_test = y_test[1:] #we shift to +1month as we need the prediction of the subsequent month 
    y_test = np.array(y_test)
    y_test = np.reshape(y_test,(len(y_test),1)) # we transform it into a 2D
    #***** X_test *****
    list_features =list(df_test.columns) # List of features 
    list_months= list(df_test.index.drop_duplicates()) # List of months in the training set 
    X_test=[]
    t_steps= 21 #Number of days (data) per month 
    count= 0    # counter
    for month in list_months:    
        features =[]  #All column features
        for i in list_features:
            feat = list(df_test_scaled.loc[month,i]) #A single column feature
            if len(feat)< t_steps: # add data from previous instance if there is not enough data (21)
                difference = t_steps - len(feat)
                feat_moins1 = list(df_test_scaled.loc[list_months[count-1],i])
                to_add= feat_moins1[-difference:]
                feat = to_add + feat
            elif len(feat)> t_steps: # truncate if there are too many data in one instance
                difference = len(feat) - t_steps
                feat = feat[difference:]
            features.append(feat)
        features = np.transpose(features)
        X_test.append(features)
        count += 1
    X_test = np.array(X_test)
    X_test = X_test[:-1,:,:] #we shift to -1month as this is the predictor
    list_months = list_months[1:] # Just to extract the dates and use them later...as the index of df_results
    return X_test, y_test, list_months  # we return list_months because are going to need them later

scaler = MinMaxScaler(feature_range = (0, 1))
training_stop_date = "2019-11"
testing_start_date ="2014-12"
macro = pd.read_csv('Macroeconomic_factors.csv') # import the macro eonomic factors
macro = macro.iloc[:,1:] # Remove dates

stocks = pd.read_csv('Stocks_names.csv') # we open the first and only path of that list # This is the list of the stocks names
stocks = list(stocks.loc[:,'Stocks'])
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
nb_rows = 5209
nb_cols =11
k = 3
s = 2
# let's obtain the column names
dates_list =[]
col_names = []

for i in range(nb_cols):
    col_names.append(sheet.cell(row=2, column=2+i).value)

# let's obtain the index names
for i in range(nb_rows):
    dt = sheet.cell(row=3+i, column=1).value
    dt = dt.strftime('%Y-%m-%d')
    dates_list.append(dt)

# path = '.\Stocks' # That's the path where all my stocks are located
df_results_summary ={} #A dictionary to store (pred vs y_test) of each stock   
df_loss = {} #A dictionary to store the LOSS score of all stocks 
for p in range(len(stocks)): #len(stocks)
    print('CA FAIT',p+1,'!!!!') # to keep track
    arrays_features = np.ones([nb_rows,nb_cols])  # this is the array where we store the features' values 
    for j in range(nb_cols):  # columns
        for i in range(nb_rows): # rows
            arrays_features[i,j] = sheet.cell(row=k+i, column=s+j).value
    df = pd.DataFrame(arrays_features, columns= col_names)
    df['Dates'] = dates_list
    s += 11
    df = pd.concat([df, macro], axis =1) # Add the macro factors to the df 
    df = Tech_Ind_Ovrnightret(df) # Add technical indicators
    XX_train, yy_train = train_inputs_targets(df)
    XX_test, yy_test,Dates_TestSet = test_inputs_targets(df)
    list_pred =[]
    list_y_test =[]
    list_of_score =[]
    for r in range(len(yy_test)): # We train dynamically de data
        print(r)
        X_train = XX_train[:178+r,:,:]
        y_train = yy_train[:178+r]
        y_test = yy_test[r]
        y_test = np.reshape(y_test,(1,len(y_test)))
        X_test = XX_test[r,:]
        X_test = np.reshape(X_test,(1,X_test.shape[0],X_test.shape[1]))
        model = Sequential()
        model.add(GRU(64, return_sequences=True, input_shape=(X_train.shape[1],X_train.shape[2]))) #the second Dim of input_shape is the number of features
        model.add(Dropout(0.2))
        model.add(GRU(8, return_sequences= False))
        model.add(Dropout(0.2))
        model.add(Dense(1, activation = 'linear'))
        model.compile(optimizer = 'adam', loss = 'mean_squared_error')
        model.fit(X_train, y_train, epochs = 120,batch_size = X_train.shape[0], verbose =0)
        score = model.evaluate(X_test, y_test)
        prediction = model.predict(X_test)
        y_test = np.array(y_test).reshape(-1) #remove 1D as it is a 2D array
        prediction = np.array(prediction).reshape(-1) #Remove 1D as it is a 2D array
        prediction = prediction.item() # to obtain the scalar value only
        y_test = y_test.item() # to obtain the scalar value only
        list_pred.append(prediction)
        list_y_test.append(y_test)
        list_of_score.append(score)
    df_loss.update({stocks[p]:list_of_score}) # df of loss scores
    df_results = pd.DataFrame(zip(list_pred,list_y_test),index=Dates_TestSet, columns=['Pred','y_test']) 
    df_results_summary.update({stocks[p]:df_results}) # dictionary of prediction + y_test for EACH stock

#create a pickle file to store the summary results (y_pred vs y)
with open('df_results_summary_GRU_strat_121.pickle', 'wb') as handle:
    pickle.dump(df_results_summary, handle, protocol=pickle.HIGHEST_PROTOCOL)
#create a pickle file to store the summary results (losses)    
with open('df_loss_per_stock_GRU_strat_121.pickle', 'wb') as handle:
    pickle.dump(df_loss, handle, protocol=pickle.HIGHEST_PROTOCOL)
