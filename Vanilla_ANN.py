import numpy as np 
import openpyxl
import pickle
import pandas as pd 
import datetime as dt
import matplotlib.pyplot as plt
np.random.seed(1000)
from keras.models import Sequential
from keras.layers import Dense
from keras.layers import Dropout
from scipy.stats import linregress
from sklearn.preprocessing import MinMaxScaler

# Create momentum function used tom compute df['Momentum']
def momentum(Close):
    returns = np.log(Close)
    x = np.arange(len(returns))
    slope, _, rvalue, _, _ = linregress(x, returns)
    return ((1 + slope) ** 252) * (rvalue ** 2)

# Adding the remaing features
def Tech_Ind_Ovrnightret(df):
    #Volatility
    
    df['daily_returns']=  df['PX_LAST'].pct_change()
    df['Volatility21d']= df['daily_returns'].rolling(21).std()
    #df = df.drop("daily_returns", axis =1)
    
    #Create 21 days Moving Average
    df['SMA21'] = df['PX_LAST'].rolling(window=21).mean()

    #Create Bollinger Bands()
    df['21sd'] = df['PX_LAST'].rolling(21).std()
    df['upper_band'] = df['SMA21'] + (df['21sd']*2)
    df['lower_band'] = df['SMA21'] - (df['21sd']*2)
    df = df.drop("21sd", axis =1)

    #Momentum
    df['Momentum'] = df['PX_LAST'].rolling(21).apply(momentum, raw=False)

    #Overnight returns
    df["Overnight_Return"] = df['PX_OPEN']/df['PX_LAST'].shift(periods=1)-1

    return df.loc[21:,:] # we start from the 21st row because there is no value for the first 21 rows of Vol 21d, SMA21, upper_band, lower_band 

#Monthly Returns
def monthly_change(feature):
    return feature.iloc[-1] / feature.iloc[0] - 1

def train_test_inputs_targets(df): 
        df['month'] = df['Dates'].astype(str).str[:7]
        df.index = df['month'] # the months are assigned as index
        df = df.drop(['month'], axis=1) # remove the months column 
        df = df.drop(['Dates'],axis=1) # remove the dates column 
        df_train = df.loc[:training_stop_date,:]
        df_test = df.loc[testing_start_date:,:]
        #Let's prepare the targets of train and test for LSTM, GRU, Vanilla Neural Network 
        grouped_train = df_train.groupby("month")
        y_train = grouped_train['PX_LAST'].apply(monthly_change)
        y_train = y_train[1:] #we shift to +1month as we need the prediction of the next month (remove the first row)
        y_train = np.array(y_train) # just converting it into numpy array
        y_train = np.reshape(y_train,(len(y_train),1)) # make it 2D for the scaler
        grouped_test = df_test.groupby("month")
        y_test = grouped_test['PX_LAST'].apply(monthly_change)
        y_test = y_test[1:] #we shift to +1month as we need the prediction of the subsequent month (remove the first row)
        Dates_list = list(y_test.index) # Just to extract the dates and use them later...as the index of df_results
        y_test = np.array(y_test) #just converting it into numpy array
        y_test = np.reshape(y_test,(len(y_test),1)) # make it 2D for the scaler
        #Let's prepare the inputs of train and test for Vanilla Neural Network 
        X_train = grouped_train.tail(1)
        X_train = X_train[:-1] #we shift to -1month as this is the predictor (remove the last row)
        inputs_train_scaled = scaler.fit_transform(X_train)
        X_test = grouped_test.tail(1)
        X_test = X_test[:-1]   #we shift to -1month as this is the predictor (remove the last row)
        inputs_test_scaled = scaler.fit_transform(X_test)
       
        return inputs_train_scaled, y_train, inputs_test_scaled, y_test,Dates_list # check what gives better results unscaled y's or scales y's 


scaler = MinMaxScaler(feature_range = (0, 1))
training_stop_date = "2019-11"
testing_start_date ="2014-12"
macro = pd.read_csv('Macroeconomic_factors.csv') # import the macro economic factors
macro = macro.iloc[:,1:] # Remove dates

stocks = pd.read_csv('Stocks_names.csv') # we open the first and only path of that list # This is the list of the stocks names
stocks = list(stocks.loc[:,'Stocks'])

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active
nb_rows = 5209
nb_cols =11
k = 3
s = 2
# let's obtain the column names
dates_list =[]
col_names = []
for i in range(nb_cols):
    col_names.append(sheet.cell(row=2, column=2+i).value)

# let's obtain the index names
for i in range(nb_rows):
    dt = sheet.cell(row=3+i, column=1).value
    dt = dt.strftime('%Y-%m-%d')
    dates_list.append(dt)

df_results_summary ={} #A dictionary to store (pred vs y_test) of each stock   
df_loss = {} #A dictionary to store the LOSS score of all stocks 
for p in range(60,80): #len(stocks)
    print('CA FAIT',p+1,'Stock !!!!')  # to keep track
    arrays_features = np.ones([nb_rows,nb_cols])  # this is the array where we store the features' values
    # stock_name = sheet.cell(row=1, column=s).value # the stock name 
    for j in range(nb_cols):  # columns
        for i in range(nb_rows): # rows
            arrays_features[i,j] = sheet.cell(row=k+i, column=s+j).value
    df = pd.DataFrame(arrays_features, columns= col_names)
    df['Dates'] = dates_list
    s += 11
    df = pd.concat([df, macro], axis =1) # Add the macro factors to the df 
    df = Tech_Ind_Ovrnightret(df) # Add technical indicators
    XX_train, yy_train, XX_test, yy_test,Dates_TestSet = train_test_inputs_targets(df) # To generate training and targets X and y
    list_pred =[]
    list_y_test =[]
    list_of_score =[]
    for r in range(len(yy_test)): # We train dynamically de data
        X_train = XX_train[:178+r,:]
        y_train = yy_train[:178+r]
        y_test = yy_test[r]
        y_test = np.reshape(y_test,(1,len(y_test)))
        X_test = XX_test[r,:]
        X_test = np.reshape(X_test,(1,len(X_test)))
        # NOW WE BUILD THE MODEL WITH THE CHOSEN PARAMETERS ####
        model = Sequential()
        model.add(Dense(64, input_dim=X_train.shape[1], activation='relu'))
        model.add(Dropout(0.02))
        model.add(Dense(8, activation='relu'))
        model.add(Dropout(0.02))
        model.add(Dense(1, activation='linear'))
        # compile the keras model
        model.compile(loss='mean_squared_error', optimizer='adam') 
        #Let's fit the model and make the prediction
        model.fit(X_train, y_train,epochs=120, batch_size=X_train.shape[0],verbose = 0) # 
        score = model.evaluate(X_test, y_test) 
        prediction = model.predict(X_test)
        y_test = np.array(y_test).reshape(-1) #remove 1D as it is a 2D array
        prediction = np.array(prediction).reshape(-1) #Remove 1D as it is a 2D array
        prediction = prediction.item() # to obtain the scalar value only
        y_test = y_test.item() # to obtain the scalar value only (without '[]')
        list_pred.append(prediction)
        list_y_test.append(y_test)
        list_of_score.append(score) 
    df_loss.update({stocks[p]:list_of_score}) # df of loss scores
    df_results = pd.DataFrame(zip(list_pred,list_y_test),index=Dates_TestSet, columns=['Pred','y_test']) 
    df_results_summary.update({stocks[p]:df_results}) # dictionary of prediction + y_test for EACH stock

#create a pickle file to store the summary results (y_pred vs y)
with open('df_results_summary_ANN_strat_14.pickle', 'wb') as handle:
    pickle.dump(df_results_summary, handle, protocol=pickle.HIGHEST_PROTOCOL)
#create a pickle file to store the summary results (losses) 
with open('df_loss_per_stock_ANN_strat_14.pickle', 'wb') as handle:
    pickle.dump(df_loss, handle, protocol=pickle.HIGHEST_PROTOCOL)



